from typing import List, AnyStr

import openpyxl
import json
import io


def fix_value(s: AnyStr):
    title = str(s).strip()
    if "/" in title:
        title = title.replace("/", "")
    if "[" in title:
        title = title.replace("[", "")
    if "]" in title:
        title = title.replace("[", "")
    if "." in title:
        title = title.replace(".", "")
    if "," in title:
        title = title.replace(".", "")
    if "?" in title:
        title = title.replace(".", "")
    title = title.lower()
    title = title.replace(" ", "")
    if len(title) > 20:
        title = title[:20]
    return title


# excel表格转json文件
def excel_to_json(excel_file, json_file_name, is_en):
    # 加载工作薄
    book = openpyxl.load_workbook(excel_file)
    # 获取sheet页
    sheet = book["sheet1"]
    # 行数
    max_row = sheet.max_row
    # 列数
    max_column = sheet.max_column
    print("max_row: %d, max_column: %d" % (max_row, max_column))
    # 结果，数组存储
    heads = []
    # 解析表头
    for column in range(max_column):
        # 读取的话行列是从（1，1）开始
        heads.append(sheet.cell(1, column + 1).value)
    # 遍历每一行
    if is_en:
        one_line = {}
        for row in range(max_row):
            if row == 0:
                continue
            for column in range(max_column):
                # 读取第二行开始每一个数据
                print(column)
                if column == 1:
                    k = heads[column]
                    cell = sheet.cell(row + 1, column + 1)
                    value = cell.value
                    if value == "null":
                        continue
                    title = fix_value(value)
                    one_line[title] = value
            print(one_line)
        book.close()
        # 将json保存为文件
        save_json_file(one_line, json_file_name)
    else:
        result = []
        for row in range(max_row):
            if row == 0:
                continue
            one_line = {}
            for column in range(max_column):
                # 读取第二行开始每一个数据
                k = heads[column]
                cell = sheet.cell(row + 1, column + 1)
                value = cell.value
                if value == "null":
                    continue
                one_line[k] = value
            print(one_line)
            result.append(one_line)
        book.close()
        results = {}
        for i in result:
            if i["对应英文"] == "none":
                continue
            title = str(i["对应英文"]).strip()
            title = fix_value(title)
            results[title] = i["字段"]
        # 将json保存为文件
        save_json_file(results, json_file_name)


# 将json保存为文件
def save_json_file(jd, json_file_name: AnyStr):
    file = io.open(json_file_name, 'w', encoding='utf-8')
    txt = json.dumps(jd, indent=2, ensure_ascii=False)
    file.write(txt)
    file.close()


if __name__ == '__main__':
    xlsx_filepath = "gcxz.xlsx"
    en_json_filepath = "en1.json"
    cn_json_filepath = "cn1.json"
    excel_to_json(xlsx_filepath, en_json_filepath, True)
    excel_to_json(xlsx_filepath, cn_json_filepath, False)
